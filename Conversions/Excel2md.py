# Excel to Markdown Converter
import os

import openpyxl

from Conversions.ConversionBase import FileConverter


class ExcelToMarkdownConverter(FileConverter):
    def convert(self, progress_callback):
        base_name = os.path.splitext(os.path.basename(self.file_path))[0]
        output_dir = os.path.join(self.output_dir, base_name)
        os.makedirs(output_dir, exist_ok=True)

        workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet_names = workbook.sheetnames

        total_sheets = len(sheet_names)
        for index, sheet_name in enumerate(sheet_names):
            sheet = workbook[sheet_name]

            markdown_content = f"# {sheet_name}\n\n"

            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                header = rows[0]
                markdown_content += f"| {' | '.join(str(cell) if cell is not None else '' for cell in header)} |\n"
                markdown_content += f"| {' | '.join(['---'] * len(header))} |\n"

                for row in rows[1:]:
                    row_content = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    markdown_content += f"| {row_content} |\n"

            md_file_path = os.path.join(output_dir, f'{sheet_name}.md')
            with open(md_file_path, 'w', encoding='utf-8') as md_file:
                md_file.write(markdown_content)

            progress_callback((index + 1) / total_sheets * 100)

        return f"Excel文件已转换为Markdown，保存到: {output_dir}"